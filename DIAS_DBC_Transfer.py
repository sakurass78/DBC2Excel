# -*- coding: utf-8 -*-
"""
Created on Tue Aug 24 15:05:24 2021

@author: Z0065679
"""
# 当需要打包成Exe软件时,删除此注释(PyQt5.sip)
# 平时编写代码时,将此句进行注释(PyQt5.sip)
# import PyQt5.sip
from sqlite3 import connect

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
import openpyxl
import pandas as pd
import sys
import os
import datetime
import re
# 导入需要的第三方库 Vector_dbc
import time

curpath = os.path.dirname(os.path.realpath(__file__))
#Fatherpath = os.path.abspath('D:\WorkPlace\EAST\DBC_Tool\CompareTool')
# print(Fatherpath)
print(curpath)
sys.path.append(curpath + "\\The3rdLibrary")
import vector_dbc

#import QtDesign_UI.First as First
import First
from PyQt5 import QtWidgets, QtCore
from PyQt5.QtWidgets import QApplication, QMainWindow, QComboBox, QMessageBox

import BitMap

import xlwings as xw
import shutil

TimeoutValueName = []
TimeoutTimeName = []

class PrintDbcToDoos(object):

    # 初始化整个Class调入DBC文件
    def __init__(self):

        # 创建 QAplplication 类
        self.app = QApplication(sys.argv)

        self.mainwindow = QMainWindow()
        self.ui = First.Ui_MainWindow()
        self.ui.setupUi(self.mainwindow)

        #显示UI后有4个选项卡，会根据选项卡被选中状态来切换UI
        self.ShowInputExcelOption()

        #信号与槽的设定
        # 初始化文件按钮
        self.ui.actionInput_DBC_File.triggered.connect(self.SelectFile)
        self.ui.InPutExcelFileSelect.clicked.connect(self.InPutSelectFile)
        self.ui.ReferenceFileSelect.clicked.connect(self.ReferenceSelectFile)

        # 链接Create按钮的调用函数
        self.ui.pushButton_2.clicked.connect(self.CreateDoorsButton)
        self.ui.AllDBCCreateButton.clicked.connect(self.AllDBCCreateButton)
        self.ui.InterfaceButtonCreate.clicked.connect(self.CreateInterfacesAutoButton)

        # 链接选择保存的文件路径 按钮
        self.ui.pushButton.clicked.connect(self.SelectOutFilePath)
        self.ui.Tabs.currentChanged.connect(self.ShowInputExcelOption)
        # 设置TextEdit 的字体
        # txtFont = QFont("Segoe UI")
        self.ui.textEdit.setFontFamily("Segoe UI")
        self.ui.textEdit_3.setFontFamily("Segoe UI")
        self.ExcelFilePath = ''
        self.DBCFilePath = ''
        self.Exceed32bitStorage = []

    def ShowInputExcelOption(self):
        if self.ui.Tabs.currentIndex() == 0:
            self.DBCParseAutoSheetTabShow(True)
            self.CreateDoorsToExcelTabShow(False)
            self.AllDBCMsgToExcelTabShow(False)
            self.InterfaceAutoSheetTabShow(False)
        elif self.ui.Tabs.currentIndex() == 1:
            self.DBCParseAutoSheetTabShow(False)
            self.CreateDoorsToExcelTabShow(False)
            self.InterfaceAutoSheetTabShow(False)
            self.AllDBCMsgToExcelTabShow(True)
        elif self.ui.Tabs.currentIndex() == 2:
            self.DBCParseAutoSheetTabShow(False)
            self.CreateDoorsToExcelTabShow(False)
            self.AllDBCMsgToExcelTabShow(False)
            self.InterfaceAutoSheetTabShow(True)
        else:
            self.DBCParseAutoSheetTabShow(False)
            self.CreateDoorsToExcelTabShow(True)
            self.InterfaceAutoSheetTabShow(False)
            self.AllDBCMsgToExcelTabShow(False)


    def DBCParseAutoSheetTabShow(self, isShow):
        if isShow == True:
            # 设置文件名的默认值
            NowDate = str(datetime.datetime.now().year) + str(datetime.datetime.now().month) + \
                      str(datetime.datetime.now().day)
            self.ui.lineEdit_2.setText("DBC_analysis_" + NowDate + ".xlsx")
            self.ui.label_3.show()
            self.ui.label_4.show()
            self.ShowNodeBoxOption(True)
        else:
            self.ui.label_3.hide()
            self.ui.label_4.hide()
            self.ShowNodeBoxOption(False)


    def InterfaceAutoSheetTabShow(self, isShow):
        if isShow == True:
            self.ui.lineEdit_2.setText("InterfacesAutoGen.xlsm")
            self.ui.ReferenceFilePath.show()
            self.ui.ReferenceFileSelect.show()
            self.ui.ReferenceFile.show()
            self.ui.ReferenceSheetList.show()
            self.ui.label_6.show()
            self.ui.InPutExcelFile.show()
            self.ui.InPutExcelFileSelect.show()
            self.ui.InPutExcelFilePath.show()
        else:
            self.ui.ReferenceFilePath.hide()
            self.ui.ReferenceFileSelect.hide()
            self.ui.ReferenceFile.hide()
            self.ui.ReferenceSheetList.hide()
            self.ui.label_6.hide()
            self.ui.InPutExcelFile.hide()
            self.ui.InPutExcelFileSelect.hide()
            self.ui.InPutExcelFilePath.hide()

    #UI界面中Tab CheckDBCfile的输出默认显示  这个不实现
    def CreateDoorsToExcelTabShow(self, isShow):
        if isShow == True:
            NowDate = str(datetime.datetime.now().year) + str(datetime.datetime.now().month) + \
                      str(datetime.datetime.now().day)
            self.ui.lineEdit_2.setText("CheckDBC_" + NowDate + ".xlsx")
        else:
            pass

    def ShowNodeBoxOption(self, isShow):
        if isShow == True:
            self.ui.label_5.show()
            self.ui.comboBox.show()
        else:
            self.ui.label_5.hide()
            self.ui.comboBox.hide()

    #UI界面中读取修改过的Msg与TX/RX信号的中间文件并转换成ALLDBC的文件名与路径设置
    def AllDBCMsgToExcelTabShow(self, isShow):
        if isShow == True:
            NowDate = str(datetime.datetime.now().year) + str(datetime.datetime.now().month) + \
                      str(datetime.datetime.now().day)
            self.ui.lineEdit_2.setText("AllDBCMsg_" + NowDate + ".xlsx")
            self.ui.InPutExcelFile.show()
            self.ui.InPutExcelFileSelect.show()
            self.ui.InPutExcelFilePath.show()
        else:
            self.ui.InPutExcelFile.hide()
            self.ui.InPutExcelFileSelect.hide()
            self.ui.InPutExcelFilePath.hide()



    def ListExcelSheet(self, strList: list = None):
        self.ui.ReferenceSheetList.clear()
        self.ui.ReferenceSheetList.addItems(strList)
        self.ui.ReferenceSheetList.SizeAdjustPolicy(QComboBox.AdjustToContents)




    # 用作重新对Class中用到的变量进行复位
    def InitClassVar(self):
        # 用作监控处理了多少个Message的变量
        # 重新载入DBC文件时需要对此变量进行清空
        self.QuantityProcessed = 0
        self.Exceed32bitStorage = []


    # 开始跑主UI
    def RunW(self):
        # 进入一个程序主循环,扫描窗口的所有事件
        # 并通过Exit函数确保正确退出,正确关闭所有资源
        self.mainwindow.show()
        sys.exit(self.app.exec_())


    # 设置下拉菜单的条目#用LoadingDBC文件把node节点list都输出
    def SetComboBoxList(self, strList: list = None):
        # 在添加下来菜单的条目时,应将菜单中已存在的条目清空
        self.ui.comboBox.clear()

        # 对下拉菜单中添加条目
        self.ui.comboBox.addItems(strList)
        # 设定下拉菜单的显示长度随字符长度匹配
        self.ui.comboBox.SizeAdjustPolicy(QComboBox.AdjustToContents)


    # UI选择DBC文件函数
    def SelectFile(self):
        # Sora Fu 补充try语句来表示如果没有正确读取DBC文件会报错
        #try:
        path = QtWidgets.QFileDialog.getOpenFileName(None, "Select DBC File", "./", "DBC File(*.dbc)")
        # 将文件路径放到对应的文本框中
        self.ui.label_4.setText(path[0])
        # 获取输入DBC文件的地址
        self.DBCFilePath = path[0]
        global DBCPATH
        DBCPATH = path[0]
        # 调用处理函数,返回下拉菜单应填入的内容
        temp = self.LoadingDBCFile(path[0])  # 用LoadingDBC文件把node节点list都输出
        self.SetComboBoxList(temp)  # UI界面下拉框可选node节点list

        # 重新载入DBC文件时需要对此变量进行清空
        self.InitClassVar()
        #except Exception as e:
        #    txt = "<font color = 'red' >[Error]</font>:\tA Serious Error happened with reading DBC file: " + str(e)
        #    print(txt)
        #    self.AddTextToTextEdit(txt)




    # 选择输出的文件夹路径
    def SelectOutFilePath(self):
        if self.DBCFilePath  == '':
            defaultpath = "./"
        else:
            # 默认存储的文件路径和DBC获取的路径相同
            temp = self.DBCFilePath.split('/')[:-1]
            # 将路径整合
            defaultpath = ''
            for index in temp:
                defaultpath = defaultpath + index + "/"

        path = QtWidgets.QFileDialog.getExistingDirectory(None, "Open Directory",
                                                          defaultpath,
                                                          QtWidgets.QFileDialog.ShowDirsOnly
                                                          | QtWidgets.QFileDialog.DontResolveSymlinks)

        # 将选择的文件夹路径放到文本框中
        self.ui.AllDBCOutputPath.setText(path)
        # 获取输出Excel的路径
        self.ExcelFilePath = path

    # 选择输入的Excel路径 用来生成AllDBC
    def InPutSelectFile(self):
        path = QtWidgets.QFileDialog.getOpenFileName(None, "Select Excel File", "./", "Excel File(*.xlsx | *.xlsm)")
        self.ui.InPutExcelFilePath.setText(path[0])

    def ReferenceSelectFile(self):
        path = QtWidgets.QFileDialog.getOpenFileName(None, "Select Refer Excel File", "./", "Excel File(*.xlsx | *.xlsm)")
        self.ui.ReferenceFilePath.setText(path[0])
        wb = openpyxl.load_workbook(path[0])
        listSheet = wb.sheetnames
        wb.close()
        self.ListExcelSheet(listSheet)

    # 输出AllDBC
    def AllDBCCreateButton(self):
        self.ExcelFileName = self.ui.lineEdit_2.text()
        self.InputExcelFilePath = self.ui.InPutExcelFilePath.text()
        self.AllDBCOutputPath = self.ui.AllDBCOutputPath.text()

        # if output path is not chosen
        if (len(self.AllDBCOutputPath)) == 0:
            self.AllDBCOutputPath = os.getcwd()
            #self.ui.AllDBCOutputPath.setText(self.AllDBCOutputPath)
        ExcelPathAllDBC = self.AllDBCOutputPath + "\\" + self.ExcelFileName
        txt = "<font color = 'blue' >[INFO]</font>:\t **********************Start to Write Messages**********************"
        self.AddTextToTextEditForAllDBCExport(txt)
        # Sora Fu AllDBC生成不需要通过DBC来做，只需要通过输入的excel文件
        #self.NodeNum = self.ui.comboBox.currentIndex()
        #self.NodeName = self.ui.comboBox.currentText()
        #self.OutputToAllDBCExcel(self.NodeNum, ExcelPathAllDBC)
        self.OutputToAllDBCExcel(self.InputExcelFilePath, ExcelPathAllDBC)



    def CreateInterfacesAutoButton(self):
        self.ExcelFileName = self.ui.lineEdit_2.text()
        self.InputExcelFilePath = self.ui.InPutExcelFilePath.text()

        # if output path is not chosen
        if (len(self.ExcelFilePath)) == 0:
            self.ExcelFilePath = os.getcwd()
            self.ui.lineEdit.setText(self.ExcelFilePath)
        OutputExcelPath = self.ExcelFilePath + "/" + self.ExcelFileName
        # 获取选择的Node的Index,返回值从0开始 和 名称
        # self.NodeNum = self.ui.comboBox.currentIndex()
        # self.NodeName = self.ui.comboBox.currentText()
        txt = "<font color = 'blue' >[INFO]</font>:\t **********************Start to Create Interfaces_Auto Sheet**********************"
        self.AddTextToTextEditForInterfacesAutoSheetAdd(txt)
        self.OutputAddInterfaceAutoSheet(self.InputExcelFilePath, OutputExcelPath)
        txt = "<font color = 'blue' >[INFO]</font>:\t **********************End to Create Interfaces_Auto Sheet**********************"
        self.AddTextToTextEditForInterfacesAutoSheetAdd(txt)

    def OutputAddInterfaceAutoSheet(self, InputExcelFile, OutputExcelPath):
        path = self.ui.ReferenceFilePath.text()
        global Wb
        if os.path.exists(InputExcelFile) == True:
            Wb = openpyxl.load_workbook(InputExcelFile, read_only=True, keep_vba=True)
        SheetName = Wb.sheetnames
        SheetName = [sheet for sheet in Wb.sheetnames if 'Interfaces' in sheet]
        if len(SheetName) > 1:
            SheetName.sort()
            InputSheetName = SheetName[len(SheetName) - 1]
            NewSheetName = 'Interfaces_Auto_' + str(int(InputSheetName[-1]) + 1)
        else:
            InputSheetName = 'Interfaces'
            NewSheetName = 'Interfaces_Auto_1'
        Wb.close()

        #copy an other file from original one
        shutil.copyfile(InputExcelFile, OutputExcelPath)
        #Add Interface Auto Sheet
        Wb2 = openpyxl.load_workbook(OutputExcelPath, read_only=False, keep_vba=False)
        Ws2 = Wb2[InputSheetName]
        target_sheet = Wb2.copy_worksheet(Ws2)
        target_sheet.title = NewSheetName
        Wb2.move_sheet(NewSheetName, -(len(Wb2.sheetnames)-1))
        # save the excel file.
        Wb2.save(OutputExcelPath)

        Wb2 = openpyxl.load_workbook(OutputExcelPath, read_only=False, keep_vba=True)
        ws2 = Wb2[NewSheetName]
        # set Progress bar 0
        self.AddProgressBarInterfaceAutoSheet(0, ws2.max_row)
        #read reference excel file as dataframe
        WsDf = pd.read_excel(io=path, sheet_name=self.ui.ReferenceSheetList.currentText())
        for rowIdx in range(1, ws2.max_row+1):
            NameToCheckTx = ws2.cell(row=rowIdx, column=1).value
            NameToCheckRx = ws2.cell(row=rowIdx, column=17).value
            TxDf = WsDf[WsDf['测试对比值(PC)'] == NameToCheckTx]
            if len(TxDf.index) > 0:
                #print(NameToCheckTx)
                for idx in TxDf.index:
                    if TxDf.loc[idx]['RXorTX(RX/TX)'] == "TX":
                        TxData = self.GetTxInterfaceInfo(TxDf, idx)
                        self.InterfacesAutoSheetTxRowUpdate(ws2, rowIdx, 1, TxData)
            RxDf = WsDf[WsDf['测试对比值(PC)'] == NameToCheckRx]
            if len(RxDf.index) > 0:
                print(NameToCheckRx)
                for idx in RxDf.index:
                    if RxDf.loc[idx]['RXorTX(RX/TX)'] == "RX":
                        #print(RxDf.loc[idx])
                        RxData = self.GetRxInterfaceInfo(RxDf, idx)
                        self.FillCellAtPosition(ws2, rowIdx, 2, 'ON')
                        self.FillCellAtPosition(ws2, rowIdx, 14, 'o')
                        self.InterfacesAutoSheetRxRowUpdate(ws2, rowIdx, 17, RxData)
            self.AddProgressBarInterfaceAutoSheet(rowIdx, ws2.max_row)
        Wb2.save(OutputExcelPath)
        Wb2.close()

    # Add Interface Auto Sheet with given row, col range
    def InterfacesAutoSheetTxRowUpdate(self, Ws, rowIdx, colStart, data):
        for colIdx in range(0, len(data)):
            Ws.cell(row=rowIdx, column=colStart+colIdx, value=data[colIdx])

    # Add Interface Auto Sheet with given row, col range
    def InterfacesAutoSheetRxRowUpdate(self, Ws, rowIdx, colStart, data):
        for colIdx in range(0, len(data)):
            if colIdx == 8:
                continue
            Ws.cell(row=rowIdx, column=colStart+colIdx, value=data[colIdx])

    def FillCellAtPosition(self, Ws, row, col, data):
        Ws.cell(row=row, column=col, value=data)

    def GetRxInterfaceInfo(self, df, idx):
        RxData = []
        RxData.append(df.loc[idx]['测试对比值(PC)'])
        DataType = df.loc[idx]['DBC_ValueType']
        BitSize = df.loc[idx]['DBC_SignalStorageBits']
        RxData.append(self.dataTypeInterfaces(str(DataType), str(BitSize)))
        RxData.append(df.loc[idx]['DBC详细信息'])
        RxData.append('COM')
        RxData.append(df.loc[idx]['DBC_DEFAULT'])
        RxData.append(df.loc[idx]['Factor'])
        RxData.append(df.loc[idx]['Offset'])
        RxData.append(df.loc[idx]['DBC_Unit'])
        RxData.append('None')
        RxData.append(df.loc[idx]['Update_Minimum'])
        RxData.append(df.loc[idx]['Update_Maximum'])
        RxData.append(df.loc[idx]['Minimum'])
        RxData.append(df.loc[idx]['Maximum'])
        RxData.append('variable')
        RxData.append('COM_API.h')
        return RxData


    def GetTxInterfaceInfo(self, df, idx):
        TxData = []
        TxData.append(df.loc[idx]['测试对比值(PC)'])
        TxData.append('ON')
        DataType = df.loc[idx]['DBC_ValueType']
        BitSize = df.loc[idx]['DBC_SignalStorageBits']
        TxData.append(self.dataTypeInterfaces(str(DataType), str(BitSize)))
        TxData.append(df.loc[idx]['DBC详细信息'])
        TxData.append('COM')
        TxData.append(df.loc[idx]['DBC_DEFAULT'])
        TxData.append(df.loc[idx]['Factor'])
        TxData.append(df.loc[idx]['Offset'])
        TxData.append(df.loc[idx]['DBC_Unit'])
        TxData.append(df.loc[idx]['Update_Minimum'])
        TxData.append(df.loc[idx]['Update_Maximum'])
        TxData.append(df.loc[idx]['Minimum'])
        TxData.append(df.loc[idx]['Maximum'])
        TxData.append('o')
        TxData.append('variable')
        TxData.append('COM_API.h')
        return TxData

    def dataTypeInterfaces(self, Sign='', Size=''):
        datatype = ''
        if Sign == 'Unsigned':
            datatype += 'uint'
        elif Sign == 'Signed':
            datatype += 'int'
        else:
            datatype += 'None'
        datatype += Size + '_T'
        return datatype

    #生成中间件的按钮操作李强开发
    def CreateDoorsButton(self):
        # 按键按下时,获取输出的Excel的Name
        self.ExcelFileName = self.ui.lineEdit_2.text()
        # 获取选择的Node的Index,返回值从0开始 和 名称
        self.NodeNum = self.ui.comboBox.currentIndex()
        self.NodeName = self.ui.comboBox.currentText() #获取node名字

        # if output path is not chosen
        if (len(self.ExcelFilePath)) == 0:
            self.ExcelFilePath = os.getcwd()
            self.ui.lineEdit.setText(self.ExcelFilePath)
        # 调用产生Execl的函数
        ExcelPath = self.ExcelFilePath + "/" + self.ExcelFileName

        self.readFile()
        #print(TimeoutValueName)
        self.OutputToXl(self.NodeNum, ExcelPath)

    #JUNJIE FU 索引超时时间与超市值自建函数，后续新增也可以用该函数
    def readFile(self): #add by JUNJIE FU
        """
        用于解析.dbc
        :return:
        """
        global node, allDatas, siganlList, SignalsName, messageName
        ''' 得到dbc文件的绝对路径'''
        filePath = DBCPATH
        #filePath = "C:\\Users\\Junjie Fu EAST\\Desktop\\CAN1.dbc"
        if filePath:
            print(filePath)
            f = open(filePath, "r", encoding='utf-8')  # 设置文件对象
        else:
            print("读取文件失败！")
            return 0
        """
        NodesPattern:节点
        MessagePattern：消息
        SignalPattern：信号
        """
        #JUNJIE FU 暂时不使用这些规则
        '''
        NodesPattern = re.compile(r"BU_: (.*)", re.S)
        MessagePattern = re.compile(r"BO_ (.*?) (.*?): (.*?) (.*)", re.S)
        SignalPattern = re.compile(SG_ (.*?) : (.*?)\|(.*?)@.*? \((.*?),(.*?)\) \[(.*?)\|(.*?)\] "(.*?)" (.*),
                                   re.S)
        DefaultValue = BA_ "GenSig(.*?)" SG_ (\d+) signalname (\d+);
        '''
        GenSigTimeoutValue = re.compile(r'BA_REL_ "GenSigTimeoutValue" (.*?) (.*?) (.*?) (.*?) (.*?) (.*);', re.S)
        GenSigTimeoutTime = re.compile(r'BA_REL_ "GenSigTimeoutTime" (.*?) (.*?) (.*?) (.*?) (.*?) (.*);', re.S)

        line = f.readline()
        allDatas = []
        while line:
            """ 匹配出节点 """
            #NodesSearched = re.search(NodesPattern, line.strip())
            '''匹配出超时时间与超时值'''
            TimeoutValueSearch = re.search(GenSigTimeoutValue, line.strip())
            TimeoutTimeSearch = re.search(GenSigTimeoutTime, line.strip())
            if TimeoutValueSearch:
                TimeoutValue = list(TimeoutValueSearch.groups())
                TimeoutValueName.append(TimeoutValue[4])
                TimeoutValueName.append(TimeoutValue[5])

            if TimeoutTimeSearch:
                TimeoutTime = list(TimeoutTimeSearch.groups())
                TimeoutTimeName.append(TimeoutTime[4])
                TimeoutTimeName.append(TimeoutTime[5])
            '''
            if NodesSearched:
                node = NodesSearched.group(1).split(" ")
            '''
            line = f.readline()

        f.close()  # 将文件关闭
        return 1

    # 向Log的框中输入内容李强开发:
    def AddTextToTextEdit(self, string):
        self.ui.textEdit.append(string)

    def AddTextToTextEditForAllDBCExport(self, string):
        self.ui.textEdit_3.append(string)

    def AddTextToTextEditForInterfacesAutoSheetAdd(self, string):
        self.ui.InterfaceTextEdit.append(string)

    # 载入DBC文件
    def LoadingDBCFile(self, DBCFilePath):

        # 开始遍历所有的Message
        # 先选择具体的Node
        self.db = vector_dbc.Database.load(DBCFilePath)
        # self.db = database.Database.load(DBCFilePath)

        # DBC中的所有节点复制到list变量中
        Node = self.db.nodes

        # 将所有的Node 名称组成一个List
        NodeNameList = []
        for NodeIndex in Node:
            NodeNameList.append(NodeIndex.name)
        return NodeNameList

    def ConfigCheck(self):
        # 对配置的属性进行判定,防止出现异常错误
        InputNameLen = len(self.sections['FileName']['inputname'])
        OutputNameLen = len(self.sections['FileName']['outputname'])

        ret = "E_OK"

        if InputNameLen == 0:
            print("[ERRO]:\t The InputName attribute in the Config.ini file has no content, please fill in ")
            ret = "E_NOT_OK"

        if OutputNameLen == 0:
            print("[ERRO]:\t The OutputName attribute in the Config.ini file has no content, please fill in ")
            ret = "E_NOT_OK"

        if ret == "E_NOT_OK":
            print("[WRAN]:\t The program will exit automatically .")
            input("Please Press Enter.......")
            sys.exit(0)

    # 从一堆Signals中寻找他们属于哪些Message 并添加到self.Tx/RxMessagelist
    def FindMessageFromSignal(self, Signals):

        Temp = []
        # 遍历TxSignals 寻找对应的Message
        for Sig in Signals:
            if Sig.message.name in Temp:
                pass
            else:
                Temp.append(Sig.message.name)

        return Temp

    # 将Signal中的大小端变更成具体的Intel或Motolar格式
    def TransmitByteOrder(self, Order):
        if Order == "big_endian":
            return "Motorola"
        if Order == "little_endian":
            return "Intel"

    # 计算Signal的StartBit
    def CalculateStartBit(self, Signal):
        # 首先判定大小端类型

        # Intel的模式下,StartBit的实际值 和 DBC文件中一致
        # Motolar的模式下,StartBit的实际值需要将DBC文件中的内容进行转换
        typeTemp = self.TransmitByteOrder(Signal.byte_order)

        if "Motorola" == typeTemp:
            # Motorla 模式
            Start = Signal.start
            Length = Signal.length
            t = Start // 8
            result = (((16 * t + Length + 6 - Start) // 8) - t) * 16 - Length + 1 + Start
            # -------------------#
            # if( Signal.name == "YawRateInvalidData" ):
            # print(" \n Motorola type")
            # print(" Start: {} ".format(Start))
            # print(" Length: {}".format(Length))
            # print(" t: {} ".format(t))
            # print(" result: {} ".format(result))

            # -------------------#

            return result

        if "Intel" == typeTemp:
            # Intel模式
            # -------------------#
            # if( Signal.name == "YawRateInvalidData" ):
            # print(" \n Intel type")

            # -------------------#
            return Signal.start

    # 开始输出一个Message的相关信息 已修改 SoraFu
    def GetMessageInfo(self, MessageName, TxRxType, Msgnum):
        Message = self.db.get_message(MessageName)
        # 获取 Node的名称 可能存在多个Node
        data = []

        #for node in Message.senders:
        #    txt = txt + str(node.name)
        #txt = txt + "\n"

        # Switch
        #data.append(Msgnum + 1)
        data.append("")
        # 获取Message的ID
        data.append(str(hex(Message.dbc_frame_id)).upper().replace("X", "x"))
        # 获取Message的Name
        data.append(Message.name)
        # 获取Message的发送或接收周期
        data.append(Message.gen_msg_cycle_time)
        # 获取Message的DLC长度
        data.append(Message.length)
        # 获取Message的format
        data.append(self.TransmitByteOrder(Message.signals[0].byte_order))
        # 获取Message的发送或接收模式
        try:
            data.append(TxRxType)
            #data.append(Message.gen_msg_send_type)
        except TypeError:
            data.append("None")
        data.append("")
        data.append("")
        data.append("")
        data.append("")
        return data


    # 获取Signal中的相关信息 已修改 SoraFu
    def GetSignalInfo(self, SignalClass, TxRxType):
        data = []
        # 获取相关报文的信息
        data.append(SignalClass.message.name)
        data.append(hex(SignalClass.message.dbc_frame_id))
        data.append(SignalClass.message.gen_msg_cycle_time)
        data.append(SignalClass.message.length)
        # 获取Signal的Name
        data.append(SignalClass.name)
        # 获取Signal的StartBit
        # 由于Motloar 和Intel 的StartBit并不一样,因此需要进行转换

        if SignalClass.name == 'AutoBrakeReqValue':
            print('find')
        temp = self.CalculateStartBit(SignalClass)

        # --------------------------#
        # print("\n Class Signal.start :{}".format(SignalClass.start))
        # print("\n startbit: {}".format(temp))
        # --------------------------#

        # 获取信号介绍描述
        # 获取Signal的Comment注释
        try:

            data.append(SignalClass.comment)
        except TypeError:
            data.append("*NO Content*")

        tempbitmap=[7, 6, 5, 4, 3, 2, 1, 0,
                   15,14,13,12,11,10, 9, 8,
                   23,22,21,20,19,18,17,16,
                   31,30,29,28,27,26,25,24,
                   39,38,37,36,35,34,33,32,
                   47,46,45,44,43,42,41,40,
                   55,54,53,52,51,50,49,48,
                   63,62,61,60,59,58,57,56]
        if SignalClass.byte_order == "Motorola":
            # 获取Signal的大小端
            #data.append(self.TransmitByteOrder(SignalClass.byte_order))

            #把startbit从LSB替换到MSB
            # Quorient 商
            Quorient = int(temp / 8)
            # remainder 余数
            Remainder = temp % 8
            Pos = Quorient * 8 + 7 - Remainder

            # 获取信号起始位
            data.append(tempbitmap[Pos-SignalClass.length+1])
        else:  #intel
            # 获取信号起始位
            data.append(SignalClass.start)

        # 获取Signal的Length
        data.append(SignalClass.length)
        # 获取Signal的数据类型
        temp = "Signed" if SignalClass.is_signed else "Unsigned"
        data.append(temp)


        # 获取Signal的factor
        data.append(SignalClass.scale)
        # 获取Signal的Offset
        data.append(SignalClass.offset)
        # 获取Signal的物理单位
        data.append(SignalClass.unit)
        # 获取Signal的最小值(物理)
        data.append(SignalClass.minimum)
        # 获取Signal的最大值(物理)
        data.append(SignalClass.maximum)
        # 计算更新factor
        data.append(1)
        # 计算更新offset
        data.append(0)
        # 计算更新最小值
        temp=(SignalClass.minimum - SignalClass.offset)/SignalClass.scale
        data.append(temp)
        # 计算更新最大值
        temp = (SignalClass.maximum - SignalClass.offset) / SignalClass.scale
        data.append(temp)
        # 获取Signal的Inti_Value:
        temp = SignalClass.gen_sig_start_value * SignalClass.scale + SignalClass.offset
        data.append(temp)
        # 获取Signal的发送类型
        #try:
        #    txt = txt + "Type:\t" + str(SignalClass.message.gen_msg_send_type) + "\n"
        #except TypeError:
        #    txt = txt + "Type:\t" + "*NO Content*" + "\n"

        # 获取default value  JUNJIE FU
        temp = SignalClass.gen_sig_start_value
        data.append(temp)
        '''
        # 获取Signal的对数值的描述
        data.append("")
        '''

        #Conversation 内容为空
        try:
            # txt = str(dict(SignalClass.choices))[1:-1]
            txt = self.outputmultilinefrom_dict_to_excel(dict(SignalClass.choices))
            data.append(txt)
        except TypeError:
            data.append("")

        #获取TimeoutValue      JUNJIEFU
        try:
            c = int(len(TimeoutValueName)/2)-1
            for i in range(0, int(len(TimeoutValueName)/2)):
                if TxRxType == "Tx":
                    data.append("")
                    break
                elif TxRxType == "Rx":
                    if SignalClass.name == TimeoutValueName[2*i]:
                        #print(TimeoutValueName[2*i+1])
                        b = TimeoutValueName[2*i+1]
                        data.append(b)
                        break
                    if i == c:
                        data.append("NONE")
                        break
                else:
                    data.append("")
        except:
           data.append("")

        #获取TimeoutTime
        try:
            d = int(len(TimeoutTimeName)/2)-1
            for i in range(0, int(len(TimeoutTimeName)/2)):
                if TxRxType == "Tx":
                    data.append("")
                    break
                elif TxRxType == "Rx":
                    if SignalClass.name == TimeoutTimeName[2*i]:
                        #print(TimeoutTimeName[2*i+1])
                        b = TimeoutTimeName[2*i+1]
                        data.append(b)
                        break
                    if i == d:
                        data.append("NONE")
                        break
                else:
                    data.append("")
        except:
           data.append("")

        return data

    def outputmultilinefrom_dict_to_excel(self, d0):
        d1 = dict(sorted(d0.items(), key=lambda item: item[0]))
        txt = ""
        for d in d1:
            txt += "" + str(d) + ": " + str(d1[d]) + "\n"
        txt = txt[:-1]
        return txt


    def sortedDictValues(self, adict):
        keys = adict.keys()
        keys.sort()
        return [adict[key] for key in keys]


    # 输出一个Message中Message的信息 和 所包含的Signals的信息
    def OutputAllInfoInMessage(self, MessageName, TxRxType, Signum, Msgnum):
        Message = self.db.get_message(MessageName)

        # 获取Message的Info
        txt = self.GetMessageInfo(MessageName, TxRxType, Msgnum)

        # 获取Message中包含的signal的信息
        if TxRxType == "Tx":
            TxRxSignals = self.TxSignals
            for sig in TxRxSignals:
                if sig in Message.signals:
                    # 当前的Signals存在Message中
                    temp = self.GetSignalInfo(sig, TxRxType)
                    self.TxSigdf.loc[Signum] = temp
                    Signum = Signum + 1
        else:
            #Sora Fu变更： Rx信号如果完全根据DBC来会漏掉不是该节点接收的报文中其他信号的数据
            #对于要计算E2E来说会有问题
            '''
            TxRxSignals = self.RxSignals
            for sig in TxRxSignals:
                if sig in Message.signals:
                    # 当前的Signals存在Message中
                    temp = self.GetSignalInfo(sig)
                    self.RxSigdf.loc[Signum] = temp
                    Signum = Signum+1
            '''
            TxRxSignals = self.RxSignals
            for sig in TxRxSignals:
                if sig in Message.signals:
                    for sig1 in Message.signals:
                        # 当前的Signals存在Message中
                        temp = self.GetSignalInfo(sig1, TxRxType)
                        self.RxSigdf.loc[Signum] = temp
                        Signum = Signum + 1
                    break
        # 输出所有的信息
        #for a in txt:
        #   print(a)
        return txt, Signum

    # 更新进度条的进度
    def AddPorcessBar(self):

        # 获取所有的Tx和Rx个数
        TxRxNum = len(self.TxMessageNameList + self.RxMessageNameList)
        setvalue = (100 * self.QuantityProcessed) / TxRxNum

        if (self.QuantityProcessed >= TxRxNum):
            setvalue = 100

        self.ui.progressBar.setValue(int(setvalue))

    def AddProgressBarForAllDBCExport(self, val, sum):
        setValue = (100 * val) / sum
        if (val >= sum):
            setValue = 100
        self.ui.AllDBCprogressBar.setValue(int(setValue))

    def AddProgressBarInterfaceAutoSheet(self, val, sum):
        setValue = (100 * val) / sum
        if (val >= sum):
            setValue = 100
        self.ui.InterfacesProgressBar_3.setValue(int(setValue))

    def CreateCANMatrixDataFrameType(self, TRType, Msgnum, Signum):

        if "Tx" == TRType:
            MessageNameListTemp = self.TxMessageNameList  # TxMessageNameList中已经赋值了Rx报文名
        else:
            MessageNameListTemp = self.RxMessageNameList #TxMessageNameList中已经赋值了Rx报文名

        # 输出提示信息
        txt = "<font color = 'blue' >[INFO]</font>:\t **********************Start to Write {} Messages**********************".format(
            TRType)
        self.AddTextToTextEdit(txt)

        # 开始遍历每个Message
        for msg in MessageNameListTemp:

            txt, tempnum = self.OutputAllInfoInMessage(msg, TRType, Signum,Msgnum) #根据MessageID读取Message的内容
            Signum = tempnum
            # 将具体的内容放入DF中
            self.Msgdf.loc[Msgnum] = txt
            Msgnum = Msgnum+1

            # 输出提示信息
            txttemp = "<font color = 'green'>[DONE]</font>:\t{}".format(msg)
            self.AddTextToTextEdit(txttemp)

            # 此时完成一个Message的写入,将进度条进行增加
            print(self.QuantityProcessed)
            self.QuantityProcessed = self.QuantityProcessed + 1
            self.AddPorcessBar()
        return Msgnum

    # 输出内容到Excel文件 SoraFu
    def OutputToXl(self, NodeNum, ExcelPath):
        self.TxSignals = self.db.nodes[NodeNum].tx_signals
        self.RxSignals = self.db.nodes[NodeNum].rx_signals

        # 获取发送和接收的Message名单
        self.TxMessageNameList = self.FindMessageFromSignal(self.TxSignals)
        self.RxMessageNameList = self.FindMessageFromSignal(self.RxSignals)

        global thin_border
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # 先建立 dataFrame的列属性 内容
        self.TxSigdf = pd.DataFrame(columns=["Message Name", "CAN ID", "Cyclic Time", "Msg Len", "Signal", "Description",
                                             "StartBit", "Length", "ValueType", "Factor", "Offset", "Unit",
                                             "Minimum Physical", "Maximum Physical", "Update_Factor",
                                             "Update_Offset", "Update_Minimum", "Update_Maximum",
                                             "Init Value", "Default Value", "Conversation", "GenSigTimeoutValue", "GenSigTimeoutTime"
                                             ])
        self.RxSigdf = pd.DataFrame(columns=["Message Name", "CAN ID", "Cyclic Time", "Msg Len", "Signal", "Description",
                                             "StartBit", "Length", "ValueType", "Factor", "Offset", "Unit",
                                             "Minimum Physical", "Maximum Physical", "Update_Factor",
                                             "Update_Offset", "Update_Minimum", "Update_Maximum",
                                             "Init Value", "Default Value", "Conversation", "GenSigTimeoutValue", "GenSigTimeoutTime"
                                             ])
        self.Msgdf = pd.DataFrame(columns=["Switch", "CAN ID", "Message Name", "Cyclic Time", "Msg Len",
                                           "Format", "Type", "Frame", "Differences", "Comment", "Link"])
        self.Messagesdf = pd.DataFrame(columns=["ALLDBC Switch", "Signal Switch", "ID", "CAN ID", "Message Name",
                                           "Cyclic Time", "Msg Len", "Format", "Type", "Frame", "Differences", "Comment", "Link"])
        ControlSheetExcelContent={'Control Data':['RX_Message_Start_Num','RX_Message_End_Num','TX_Message_Start_Num',
                                                   'TX_Message_End_Num','Develop_Message_Start_Num',
                                                   'Develop_Message_End_Num'],
                                  'Messsage ID':['','','','','',''],
                                  'Operation_Control':["DBC_To_Excel_Message","DBC_To_Excel_Tx_Signal",
                                                       "DBC_To_Excel_Rx_Signal","Excel_To_Matrix", '', ''],
                                  'Switch':["On","On","On","On",'','']}
        self.Controldf = pd.DataFrame(ControlSheetExcelContent)

        self.RxSignaldf = pd.DataFrame(columns=["Message Name", "CAN ID", "Cyclic Time", "Msg Len ", "Signal", "Description",
                                                "StartBit", "Length", "ValueType", "Factor", "Offset", "Unit", "Minimum Physical",
                                                "Maximum Physical", "Update_Factor", "Update_Offset", "Update_Minimum",
                                                "Update_Maximum", "Init Value", "Default Value", "Conversation"
                                                ])
        self.TxSignaldf = pd.DataFrame(columns=["Message Name", "CAN ID", "Cyclic Time", "Msg Len ", "Signal", "Description",
                                                "StartBit", "Length", "ValueType", "Factor", "Offset", "Unit", "Minimum Physical",
                                                "Maximum Physical", "Update_Factor", "Update_Offset", "Update_Minimum",
                                                "Update_Maximum", "Init Value", "Default Value", "Conversation"
                                                ])
        self.DevelopSignalsdf = pd.DataFrame(columns=["CANID(Expand)","MessageName(Expand)","DBC_Signalattribute",
                                                "DBC_SignalName","From DBC",
                                                "ControllerX(0/1/2)","RXorTX(RX/TX)","StartBit","Length",
                                                "DBC_SignalStorageBits","DBC_ValueType","Factor","Offset","DBC_INIT",
                                                "DBC_DEFAULT","DBC_Unit","MotorORIntel","Minimum","Maximum","DBC详细信息",
                                                "APP详细信息","APP_SignalName"])

        self.CheckResultdf = pd.DataFrame(columns=["No", "对比结果", "日期"])

        TxSignum = 0
        RxSignum = 0
        Msgnum = 0
        # Reset Progress bar
        self.QuantityProcessed = 0
        # 生成TX的DataFrame信息
        tempMsgnum = self.CreateCANMatrixDataFrameType("Rx",Msgnum,TxSignum)
        Msgnum = tempMsgnum
        # 生成RX的DataFrame的信息
        tempMsgnum = self.CreateCANMatrixDataFrameType("Tx",Msgnum,RxSignum)
        Msgnum = tempMsgnum
        # 写入CSV文件中
        #self.Msgdf.to_csv(ExcelPath, encoding='utf_8_sig', index=False)
        #self.TxSigdf.to_csv(ExcelPathTx, encoding='utf_8_sig', index=False)
        #self.RxSigdf.to_csv(ExcelPathRx, encoding='utf_8_sig', index=False)
        MsgDfTemp=self.Msgdf
        MsgDfTemp.loc[Msgnum] = ""  # add one space row between cyclic none 0 and cyclic 0
        self.Msgdf = self.DataFrameCyclicSort(MsgDfTemp, Msgnum)
        self.GeneratorControlSheet(ExcelPath, wsname='Control')
        self.GeneratorMessagesSheet(ExcelPath, wsname="Messages")
        self.GeneratorRxSignalsSheet(ExcelPath, wsname="RX_Signals")
        self.GeneratorTxSignalsSheet(ExcelPath, wsname="TX_Signals")
        self.GeneratorDevelopSignalsSheet(ExcelPath, wsname="Develop_Signals")
        self.GeneratorCheckResultSheet(ExcelPath, wsname="Check_Result")
        self.GeneratorExcel(ExcelPath, wsname='DBC_Messages')
        self.GeneratorExcel(ExcelPath, wsname='DBC_RX_Signals')
        self.GeneratorExcel(ExcelPath, wsname='DBC_TX_Signals')

    def OutputToAllDBCExcel(self, InputExcelFilePath, ExcelPathAllDBC):
        self.GreenRowIdx = 0
        #self.TxSignals = self.db.nodes[NodeNum].tx_signals
        #self.RxSignals = self.db.nodes[NodeNum].rx_signals

        # 获取发送和接收的Message名单
        #self.TxMessageNameList = self.FindMessageFromSignal(self.TxSignals)
        #self.RxMessageNameList = self.FindMessageFromSignal(self.RxSignals)

        global thin_border
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # 先建立 dataFrame的列属性 内容
        '''
        self.TxSigdf = pd.DataFrame(columns=["Message Name", "CAN ID", "Cyclic Time", "Msg Len", "Signal", "Description",
                                             "StartBit", "Length", "ValueType", "Factor", "Offset", "Unit",
                                             "Minimum Physical", "Maximum Physical", "Update_Factor",
                                             "Update_Offset", "Update_Minimum", "Update_Maximum",
                                             "Init Value", "Default Value", "Conversation"])
        self.RxSigdf = pd.DataFrame(columns=["Message Name", "CAN ID", "Cyclic Time", "Msg Len", "Signal", "Description",
                                             "StartBit", "Length", "ValueType", "Factor", "Offset", "Unit",
                                             "Minimum Physical", "Maximum Physical", "Update_Factor",
                                             "Update_Offset", "Update_Minimum", "Update_Maximum",
                                             "Init Value", "Default Value", "Conversation"])
        self.Msgdf = pd.DataFrame(columns=["Switch", "CAN ID", "Message Name", "Cyclic Time", "Msg Len",
                                           "Format", "Type", "Frame", "Differences", "Comment", "Link"])
        '''
        self.AllDBCMsgBluePartdf = pd.DataFrame(columns=["CANID(Total)","MessageName(Total)","FRAME_TYPE",
                                                             "DIAG_TimerOut","DIAG_RC","DIAG_CS","DIAG_SV","RollCount",
                                                             "RX(Node)orTX(Node)","MessageLength","GenMsgCycleTime"])
        self.AllDBCMsgGreenPartdf= pd.DataFrame(columns=["CANID(Expand)","MessageName(Expand)","DBC_Signalattribute",
                                                             "DBC_SignalName","From DBC","ControllerX(0/1/2)","RXorTX(RX/TX)",
                                                             "StartBit","Length","DBC_SignalStorageBits","DBC_ValueType",
                                                             "Factor","Offset","DBC_INIT","DBC_DEFAULT","DBC_Unit",
                                                             "MotorORIntel","Minimum","Maximum","DBC详细信息", "APP详细信息",
                                                             "APP_SignalName", "ASW_SignalValid", "APP_Factor",
                                                             "APP_Offset", "APP_SignalStorageBits", "APP_Default Value",
                                                             "APP_Fault Value", "APP_Unit", "ENUM_APP", "ENUM_DBC",
                                                             "Calc_Factor", "Update_Factor", "Update_Offset",
                                                             "Update_Minimum", "Update_Maximum", "Update_APP_Factor",
                                                             "Update_APP_Offset", "测试对比值(PC)", "测试输出值(ECU)", "测试结果"])

        TxSignum = 0
        RxSignum = 0
        Msgnum = 0
        # 生成TX的DataFrame信息
        #tempMsgnum = self.CreateCANMatrixDataFrameType("Rx",Msgnum,TxSignum)
        #Msgnum = tempMsgnum
        # 生成RX的DataFrame的信息
        #tempMsgnum = self.CreateCANMatrixDataFrameType("Tx",Msgnum,RxSignum)
        #Msgnum = tempMsgnum
        #self.DBCRxSigdf = self.RxSigdf
        #self.DBCTxSigdf = self.TxSigdf
        self.StartGreenPartColPos = len(self.AllDBCMsgBluePartdf.columns)
        #MsgDfTemp=self.Msgdf
        #MsgDfTemp.loc[Msgnum] = ""  # add one space row between cyclic none 0 and cyclic 0
        #self.Msgdf = self.DataFrameCyclicSort(MsgDfTemp, Msgnum)

        self.WriteTitleToExcel(OutPath=ExcelPathAllDBC, wsname='AllDBCMessage_handle_NEW_BASE_S',
                               df=self.AllDBCMsgBluePartdf, colPos=0,  isClearSheet=True)
        self.WriteTitleToExcel(ExcelPathAllDBC, 'AllDBCMessage_handle_NEW_BASE_S', self.AllDBCMsgGreenPartdf,
                               colPos=self.StartGreenPartColPos, isClearSheet=False)

        TxSigDf = self.ReadOutputSheetAsDataFrame(self.InputExcelFilePath, 'TX_Signals')
        RxSigDf = self.ReadOutputSheetAsDataFrame(self.InputExcelFilePath, 'RX_Signals')
        WsDf = self.ReadOutputSheetAsDataFrame(self.InputExcelFilePath, 'Messages')
        for Idx in WsDf.index:
            data = self.GetDataForBluePart(WsDf, Idx)
            self.GetDataForGreenPart(WsDf, Idx, TxSigDf, RxSigDf)
            if not data:
                self.AllDBCMsgBluePartdf.loc[Idx] = ""
            elif data[0] == 'OFF':
                continue
            else:
                self.AllDBCMsgBluePartdf.loc[Idx] = data
        self.GeneratorAllDBCMessages(WsDf, ExcelPathAllDBC, wsname='AllDBCMessage_handle_NEW_BASE_S')

        # 往AllDBC中写入数据

    def GeneratorAllDBCMessages(self, WsDf, OutPath='', wsname=''):
        isWsExist = False
        isClear = False
        FillBlue = PatternFill('solid', fgColor='33CCFF')
        FillGreen = PatternFill('solid', fgColor='33CC66')
        FillYellow = PatternFill('solid', fgColor='FFFF33')
        if os.path.exists(OutPath) == True:
            Wb = openpyxl.load_workbook(OutPath)
        else:
            Wb = Workbook()  # 创建新的excel
        for sheet in Wb.sheetnames:
            if sheet == wsname:
                Ws = Wb[wsname]
                if isClear == True:
                    Wb.remove(Wb[wsname])
                    isWsExist = False
                else:
                    isWsExist = True
                break
            else:
                continue
        if not isWsExist:
            if wsname == 'AllDBCMessage_handle_NEW_BASE_S':
                Ws = Wb.create_sheet(wsname, len(Wb.worksheets) + 1)
                # Message的格式
                Ws.row_dimensions[1].height = 40  # 行高设置，行1 40
                Ws.column_dimensions['A'].width = 12  # 列宽设置， 列A 10
        rowindex = 2
        for idx in self.AllDBCMsgBluePartdf.index:
            for Var in range(0, len(self.AllDBCMsgBluePartdf.columns)):
                Ws.cell(row=rowindex, column=Var + 1).value = self.AllDBCMsgBluePartdf.loc[idx][Var]
                Ws.cell(row=rowindex, column=Var + 1).border = thin_border
                Ws.cell(row=rowindex, column=Var + 1).font = Font(u'Arial', size=10, bold=False,
                                                                  italic=False, strike=False, color='000000')
                Ws.cell(row=rowindex, column=Var + 1).fill = FillBlue

            txt = "<font color = 'green'>[DONE]</font>:\t{}".format(
                self.AllDBCMsgBluePartdf.loc[idx]['MessageName(Total)'])
            self.AddProgressBarForAllDBCExport(rowindex - 1, len(self.AllDBCMsgBluePartdf.index))
            if len(txt) > 40:
                self.AddTextToTextEditForAllDBCExport(txt)
            rowindex += 1

        rowindex = 2
        for idx in self.AllDBCMsgGreenPartdf.index:
            for Var in range(0, len(self.AllDBCMsgGreenPartdf.columns)):
                colPos = self.StartGreenPartColPos + 1
                Ws.cell(row=rowindex, column=colPos + Var).value = self.AllDBCMsgGreenPartdf.loc[idx][Var]
                Ws.cell(row=rowindex, column=colPos + Var).border = thin_border
                if Var == 9 and idx in self.Exceed32bitStorage:
                    Ws.cell(row=rowindex, column=colPos + Var).fill = FillYellow
                else:
                    Ws.cell(row=rowindex, column=colPos + Var).fill = FillGreen
                Ws.cell(row=rowindex, column=colPos + Var).font = Font(u'Arial', size=10, bold=False,
                                                                       italic=False, strike=False, color='000000')
            rowindex += 1
        txt = "<font color = 'blue' >[INFO]</font>:\t **********************End to Write Messages**********************"
        self.AddTextToTextEditForAllDBCExport(txt)

        if sheet == "Sheet":
            Wb.remove(Wb[sheet])
        Wb.save(OutPath)
        Wb.close()

        # 获取需要输入AllDBC中的蓝色区域 也就是Tx Rx信号的数据

    #导出Message没问题
    def GetDataForBluePart(self, WsDf, Idx):
        BluePartData = []
        if WsDf.loc[Idx]['ALLDBC Switch'] == 'ON':
            # Add CANID(Total)
            BluePartData.append(str(WsDf.loc[Idx]['CAN ID']).lower())
            # Add MessageName(Total)
            BluePartData.append(WsDf.loc[Idx]['Message Name'])
            # Add Frame Type
            FrameType = self.GetFrameType(WsDf.loc[Idx]['Frame'], WsDf.loc[Idx]['Type'])
            BluePartData.append(FrameType)
            # Add DIAG_TimerOut
            BluePartData.append('N')
            # Add DIAG_RC
            BluePartData.append('N')
            # Add DIAG_CS
            BluePartData.append('N')
            # Add DIAG_SV
            BluePartData.append('N')
            # Add RollCount
            BluePartData.append(4)
            # Add RX(Node)orTX(Node)
            if WsDf.loc[Idx]['Type'] == 'Rx':
                BluePartData.append('RX0')
            else:
                BluePartData.append('TX0')
            # Add MessageLength
            BluePartData.append(WsDf.loc[Idx]['Msg Len'])
            # Add GenMsgCycleTime
            if WsDf.loc[Idx]['Cyclic Time'] == 0:
                BluePartData.append(10)
            else:
                BluePartData.append(WsDf.loc[Idx]['Cyclic Time'])
        elif WsDf.loc[Idx]['ALLDBC Switch'] == 'OFF':
            BluePartData = ['OFF']
        return BluePartData

        # 获取需要输入AllDBC中的绿色区域 也就是报文的数据

    def GetDataForGreenPart(self, WsDf, Idx, TxSigDf, RxSigDf):
        tmp = None
        if WsDf.loc[Idx]['ALLDBC Switch'] == 'ON' and WsDf.loc[Idx]['Signal Switch'] == 'ON':
            RxorTx = WsDf.loc[Idx]['Type'].upper()
            if RxorTx == 'TX':
                DBCTxRxSignalTmp = TxSigDf
            else:
                DBCTxRxSignalTmp = RxSigDf

            #从MSG中找出Idx行的canID，并要从信号表里面找出对应的信号
            MsgCANID = WsDf.loc[Idx]['CAN ID'].lower()
            #tmp = DBCTxRxSignalTmp[DBCTxRxSignalTmp['CAN ID'] == MsgCANID]
            #tmp = tmp.sort_values(by='StartBit')
            flag_row = False
            for num in range(0, len(DBCTxRxSignalTmp['CAN ID'])):
                if DBCTxRxSignalTmp.loc[num]['CAN ID'] == MsgCANID:
                    row1 = num
                    #row1 = DBCTxRxSignalTmp[DBCTxRxSignalTmp['CAN ID'].isin([row])]
                    flag_row = True
                    num += 1
                temp = DBCTxRxSignalTmp.loc[num]['Signal']
                if flag_row == True and ((num == len(DBCTxRxSignalTmp['CAN ID']) - 1) or (pd.isna(DBCTxRxSignalTmp.loc[num]['Signal']) == True)):
                    flag_row = False
                    row2 = num
                    #row2 = DBCTxRxSignalTmp[DBCTxRxSignalTmp['CAN ID'].isin([row])]
                    tmp = DBCTxRxSignalTmp[row1:row2]
                    ID = MsgCANID
                    MsgNm = tmp.loc[row1]['Message Name']

                    for idx in tmp.index:
                        if tmp.loc[idx]['CAN ID'] != MsgCANID :
                            GreenPartData = []
                            #ID = tmp.loc[idx]['CAN ID']
                            SignalName = tmp.loc[idx]['Signal']
                            ValueType = tmp.loc[idx]['ValueType']
                            storageBit = self.GetDBCSignalStorageBits(tmp, idx)

                            #添加CANID
                            GreenPartData.append(ID)
                            # 添加MSG名字
                            GreenPartData.append(MsgNm)
                            # 添加DBC_Signalattribute
                            GreenPartData.append('Value_Signal')
                            # 添加信号名
                            GreenPartData.append(SignalName)
                            # From DBC
                            GreenPartData.append(1)
                            # ControllerX(0/1/2)
                            GreenPartData.append(0)
                            # RX or Tx
                            GreenPartData.append(RxorTx)
                            # StartBit
                            GreenPartData.append(tmp.loc[idx]['StartBit'])
                            # Length
                            GreenPartData.append(tmp.loc[idx]['Length'])
                            # DBC_SignalStorageBits

                            GreenPartData.append(storageBit)
                            # Value type
                            GreenPartData.append(ValueType)
                            # Factor
                            GreenPartData.append(tmp.loc[idx]['Factor'])
                            # Offset
                            GreenPartData.append(tmp.loc[idx]['Offset'])
                            # DBC_INIT
                            GreenPartData.append(tmp.loc[idx]['Init Value'])
                            # DBC_DEFAULT
                            GreenPartData.append(tmp.loc[idx]['Default Value'])
                            # DBC_Unit
                            GreenPartData.append(tmp.loc[idx]['Unit'])
                            # MotorORIntel
                            GreenPartData.append(WsDf.loc[Idx]['Format'])
                            # Minimum
                            GreenPartData.append(tmp.loc[idx]['Minimum Physical'])
                            # Maximum
                            GreenPartData.append(tmp.loc[idx]['Maximum Physical'])
                            # DBC详细信息
                            space = '\n'
                            Conversation = tmp.loc[idx]['Conversation']
                            Description = tmp.loc[idx]['Description']
                            if pd.isna(tmp.loc[idx]['Description']) == True:
                                if pd.isna(tmp.loc[idx]['Conversation']) == True:
                                    GreenPartData.append('')
                                else:
                                    GreenPartData.append(str(Conversation).rstrip())
                            else:
                                if pd.isna(tmp.loc[idx]['Conversation']) == True:
                                    GreenPartData.append(Description)
                                else:
                                    GreenPartData.append(Description + space + str(Conversation).rstrip())
                            # APP详细信息
                            GreenPartData.append('')
                            # APP_SignalName
                            GreenPartData.append('')
                            # ASW_SignalValid
                            GreenPartData.append('')
                            # APP_Factor
                            GreenPartData.append('')
                            # APP_Offset
                            GreenPartData.append('')
                            # APP_SignalStorageBits
                            GreenPartData.append('')
                            # APP_Default Value
                            GreenPartData.append('')
                            # APP_Fault Value
                            GreenPartData.append('')
                            # APP_Unit
                            GreenPartData.append('')
                            # ENUM_APP
                            GreenPartData.append('')
                            # ENUM_DBC
                            GreenPartData.append('')
                            # Calc_Factor
                            GreenPartData.append('')
                            GreenPartData.append(tmp.loc[idx]['Update_Factor'])
                            GreenPartData.append(tmp.loc[idx]['Update_Offset'])
                            GreenPartData.append(tmp.loc[idx]['Update_Minimum'])
                            GreenPartData.append(tmp.loc[idx]['Update_Maximum'])
                            GreenPartData.append('')
                            GreenPartData.append('')
                            GreenPartData.append(self.SignalNameCombine(ValueType, storageBit, SignalName, ID))
                            GreenPartData.append('')
                            GreenPartData.append('')

                            self.AllDBCMsgGreenPartdf.loc[self.GreenRowIdx] = GreenPartData
                            self.GreenRowIdx += 1

    # add one space row between cyclic none 0 and cyclic 0
    def DataFrameCyclicSort(self, Dataframe, Msgnum):
        TmpMsgCnt = Msgnum + 1
        #Dataframe.loc[Msgnum] = "" #add one space row between cyclic none 0 and cyclic 0
        for i in range(0, Msgnum):
            if Dataframe.loc[i,"Cyclic Time"] == 0:
                Dataframe.loc[TmpMsgCnt] = Dataframe.loc[i]
                Dataframe = Dataframe.drop(i)
                TmpMsgCnt=TmpMsgCnt+1
        return Dataframe

    #生成TX RX信号的信号映射位图
    def GeneratorExcel(self, OutPath='', wsname=''):
        # Ws = Wb.active
        # Sheet1 Messages
        # Sheet2 TX Signals
        # Sheet3 RX Signals
        Logtxt = ''
        isWsExist = False
        isClear = True
        if os.path.exists(OutPath) == True:
            Wb = openpyxl.load_workbook(OutPath)
        else:
            Wb = Workbook()  # 创建新的excel

        for sheet in Wb.sheetnames:
            if sheet == wsname:
                Ws = Wb[wsname]
                rowindex = Ws.max_row + 1
                if isClear == True:
                    Wb.remove(Wb[wsname])
                    isWsExist = False
                else:
                    isWsExist = True
                break
            else:
                continue
        if not isWsExist:
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            if wsname == 'DBC_Messages':
                Ws = Wb.create_sheet(wsname, len(Wb.worksheets) + 1)
                # Message的格式
                #Ws.row_dimensions[1].height = 11  # 行高设置，行1 40
                Ws.column_dimensions['A'].width = 9  # 列宽设置， 列A 10
                Ws.column_dimensions['B'].width = 9
                Ws.column_dimensions['C'].width = 24
                Ws.column_dimensions['D'].width = 9
                Ws.column_dimensions['E'].width = 9
                Ws.column_dimensions['F'].width = 9
                Ws.column_dimensions['G'].width = 9
                Ws.column_dimensions['H'].width = 9

                Fill0 = PatternFill('solid', fgColor='0070C0') #0070C0
                for i in range(1,12):
                    Ws.cell(row=1, column=i).alignment = Alignment(wrapText=False)
                    #Ws.cell(row=1, column=i).fill = Fill0 #sangnt comment
                    Ws.cell(row=1, column=i).font = Font(u'Calibri', size=None, bold=True, italic=False, strike=False,
                                                         color='000000') #等线
                    Ws.cell(row=1, column=i).border = thin_border
                rowindex = 1
                Var = 0
                for col in self.Msgdf.columns:
                    Ws.cell(row=rowindex, column=Var + 1).value = col
                    Var += 1
                rowindex += 1
                for id in self.Msgdf.index:
                    for Var in range(0, len(self.Msgdf.columns)):
                        Ws.cell(row=rowindex, column=Var +1).value = self.Msgdf.loc[id][Var]
                    rowindex += 1
            else:  # Tx Rx信号的格式
                Ws = Wb.create_sheet(wsname, len(Wb.worksheets) + 1)

                #Ws.row_dimensions[1].height = 40  # 行高设置，行1 40 #sangnt comment
                Ws.column_dimensions['A'].width = 25  # 列宽设置， 列A 25
                Ws.column_dimensions['B'].width = 10
                Ws.column_dimensions['C'].width = 10
                Ws.column_dimensions['D'].width = 6
                Ws.column_dimensions['E'].width = 21
                Ws.column_dimensions['F'].width = 27
                Ws.column_dimensions['G'].width = 10
                Ws.column_dimensions['H'].width = 10
                Ws.column_dimensions['I'].width = 10
                Ws.column_dimensions['J'].width = 10
                Ws.column_dimensions['K'].width = 10
                Ws.column_dimensions['L'].width = 10
                Ws.column_dimensions['M'].width = 10
                Ws.column_dimensions['N'].width = 10
                Ws.column_dimensions['O'].width = 10
                Ws.column_dimensions['P'].width = 10
                Ws.column_dimensions['Q'].width = 10
                Ws.column_dimensions['P'].width = 10
                Ws.column_dimensions['Q'].width = 10
                Ws.column_dimensions['R'].width = 10
                Ws.column_dimensions['S'].width = 10
                Ws.column_dimensions['T'].width = 23

                Fill0 = PatternFill('solid', fgColor='99CCFF')
                for i in range(1, 35):
                    Ws.cell(row=1, column=i).alignment = Alignment(wrapText=False, horizontal='center')
                    Ws.cell(row=1, column=i).fill = Fill0
                    Ws.cell(row=1, column=i).font = Font(u'Calibri', size=11, bold=True, italic=False, strike=False,
                                                         color='000000')
                if wsname=='DBC_TX_Signals':
                    Sigdf = self.TxSigdf
                else:
                    Sigdf = self.RxSigdf

                rowindex = 1 #Excel 行数
                Var = 0 #标题列数
                for i in range(1, len(Sigdf.columns)+1):
                    Ws.cell(row=1, column=i).border = thin_border

                for col in Sigdf.columns: #写入标题
                    Ws.cell(row=rowindex, column=Var + 1).value = col
                    Var += 1
                rowindex += 1

                id = 0 #报文数
                while id < len(Sigdf.index):
                    tempdf = Sigdf[Sigdf["CAN ID"] == Sigdf.loc[id, "CAN ID"]]
                    tempdf = tempdf.sort_values(by="StartBit") #调取相同CAN ID并排序 升序

                    #JUNJIE !!!!!!!!! 屏蔽此次，可以将cycle time ==0的 message一起输出
                    #                       如果打开此处屏蔽 则输出不带 cycle 0 的message到 EXCEL
                    '''''
                    if tempdf.loc[id]['Cyclic Time'] == 0:  # not include cyclic 0 item
                        id += 1
                        continue
                    '''
                    if len(tempdf) != 1:
                        id += len(tempdf) - 1
                    countnum = 0 #区分报文行与信号行
                    global base_layout_column
                    base_layout_column = 25 #用于记录layout的基准列，最左列
                    for id2 in tempdf.index: #id2为一个报文中的信号数
                        if countnum == 0:
                            for Var in range(0, 4):
                                Ws.cell(row=rowindex, column=Var + 1).value = tempdf.loc[id2][Var]
                                Ws.cell(row=rowindex, column=Var + 1).fill = PatternFill('solid', fgColor='CCFFFF')
                            rowindex += 1
                            countnum += 1
                            BitRow = rowindex
                            for Var in range(4, len(tempdf.columns)):
                                Ws.cell(row=rowindex, column=Var + 1).value = tempdf.loc[id2][Var]
                                Ws.cell(row=rowindex, column=Var + 1).fill = PatternFill('solid', fgColor='CCFFFF')
                            rowindex += 1

                            #对信号排布进行画图
                            # BitM = BitMap.BitMap(tempdf)
                            # BitList = BitM.BitMapping()
                            # BitRow2=BitRow
                            # #写入BitMap
                            #
                            # Ws.cell(row=BitRow2, column=base_layout_column + 1).value = "位顺序"
                            # Ws.cell(row=BitRow2, column=base_layout_column + 2).value = "MSB"
                            # Ws.cell(row=BitRow2, column=base_layout_column + 9).value = "LSB"
                            # for j in range(base_layout_column, base_layout_column + 10):
                            #     Ws.cell(row=BitRow2, column=j).fill = PatternFill('solid', fgColor='FFFFCC')
                            # BitRow2 += 1
                            #
                            # Ws.cell(row=BitRow2, column=base_layout_column + 1).value = "bit order"
                            # for j in range(base_layout_column, base_layout_column + 10):
                            #     Ws.cell(row=BitRow2, column=j).fill = PatternFill('solid', fgColor='FFFFCC')
                            # BitRow2 += 1
                            #
                            # Ws.cell(row=BitRow2, column=base_layout_column).value = "字节顺序"
                            # for j in range(base_layout_column, base_layout_column + 10):
                            #     Ws.cell(row=BitRow2, column=j).fill = PatternFill('solid', fgColor='FFFFCC')
                            # for j in range(0, 8):
                            #     Ws.cell(row=BitRow2, column=base_layout_column + 2 + j).value = "bit" + str(7-j)
                            #     Ws.cell(row=BitRow2, column=base_layout_column + 2 + j).fill = PatternFill('solid', fgColor='FFFFCC')
                            # BitRow2 += 1
                            #
                            # Ws.cell(row=BitRow2, column=base_layout_column).value = "Byte Order"
                            # for j in range(base_layout_column, base_layout_column + 10):
                            #     Ws.cell(row=BitRow2, column=base_layout_column).fill = PatternFill('solid', fgColor='FFFFCC')
                            # BitRow2 += 1
                            #
                            # Ws.cell(row=BitRow2, column=base_layout_column).value = "MSB"
                            # Ws.cell(row=BitRow2+7, column=base_layout_column).value = "LSB"
                            # for k in range(0,8):
                            #     Ws.cell(row=BitRow2+k, column=base_layout_column).fill = PatternFill('solid', fgColor='FFFFCC')
                            # for j in range(0,8):
                            #     Ws.cell(row=BitRow2, column=base_layout_column + 1).value = "Byte" + str(j)
                            #     Ws.cell(row=BitRow2, column=base_layout_column + 1).fill = PatternFill('solid', fgColor='FFFFCC')
                            #     for i in range(0,8):
                            #         Ws.cell(row=BitRow2, column=base_layout_column + 2 + i).value = BitList[0][j*8+i]
                            #         Ws.cell(row=BitRow2, column=base_layout_column + 2 + i).fill = PatternFill('solid', fgColor=BitList[1][j*8+i])
                            #
                            #     BitRow2 += 1

                        else:
                            for Var in range(4, len(tempdf.columns)):
                                Ws.cell(row=rowindex, column=Var +1).value = tempdf.loc[id2][Var]
                                Ws.cell(row=rowindex, column=Var + 1).fill = PatternFill('solid', fgColor='CCFFFF')
                            rowindex += 1

                    if len(tempdf.index) > 12:  # 如果信号数大于12行
                        #报文之间空行涂灰
                        for Var in range(0,len(tempdf.columns)):
                            Ws.cell(row=rowindex, column=Var + 1).fill = PatternFill('solid', fgColor='D9D9D9')
                        rowindex += 1
                    else:
                        rowindex = BitRow + 12
                        for Var in range(0,len(tempdf.columns)):
                            Ws.cell(row=rowindex, column=Var + 1).fill = PatternFill('solid', fgColor='D9D9D9')
                        rowindex += 1
                    id += 1
                Ws.merge_cells(start_row=1, end_row=1, start_column=base_layout_column, end_column=base_layout_column+9)
                Ws.cell(row=1, column=base_layout_column).value = "Signal Layout"
                Ws.cell(row=1, column=base_layout_column).alignment = Alignment(horizontal='center')
                #Ws.cell(row=1, column=base_layout_column).border = thin_border
            if sheet == "Sheet":
                Wb.remove(Wb[sheet])

        Wb.save(OutPath)
        Wb.close()
        #JUNJIE FU  增加打印 提示输出 DBC RX/TX Signal sheet结束
        txt = "<font color = 'yellow' >[INFO]</font>:\t **********************Over Write Tx&Rx Messages**********************"
        self.AddTextToTextEdit(txt)

    def GeneratorControlSheet(self, OutPath='', wsname=''):
        isWsExist = False
        isClear = True
        if os.path.exists(OutPath) == True:
            Wb = openpyxl.load_workbook(OutPath)
        else:
            Wb = Workbook()  # 创建新的excel

        for sheet in Wb.sheetnames:
            if sheet == wsname:
                Ws = Wb[wsname]
                rowindex = Ws.max_row + 1
                if isClear == True:
                    Wb.remove(Wb[wsname])
                    isWsExist = False
                else:
                    isWsExist = True
                break
            else:
                continue

        if not isWsExist:
            if wsname == 'Control':
                Ws = Wb.create_sheet(wsname, len(Wb.worksheets) + 1)
                # Message的格式
                #Ws.row_dimensions[1].height = 11  # 行高设置，行1 40
                Ws.column_dimensions['A'].width = 28  # 列宽设置， 列A 10
                Ws.column_dimensions['B'].width = 14
                Ws.column_dimensions['C'].width = 28
                Ws.column_dimensions['D'].width = 12
                Ws.column_dimensions['E'].width = 9
                Ws.column_dimensions['F'].width = 9
                Ws.column_dimensions['G'].width = 9
                Ws.column_dimensions['H'].width = 9
            rowindex=1
            Var=1
            for Title in self.Controldf.columns:
                Ws.cell(row=rowindex, column=Var).value = Title
                Ws.cell(row=rowindex, column=Var).border = thin_border
                Ws.cell(row=rowindex, column=Var).font = Font(bold=True)
                Var+=1
            rowindex += 1
            for id in self.Controldf.index:
                for Var in range(0, len(self.Controldf.columns)):
                    Ws.cell(row=rowindex, column=Var + 1).value = self.Controldf.loc[id][Var]
                    if (id == 4 or id == 5) and (Var == 2 or Var == 3):
                        continue
                    else:
                        Ws.cell(row=rowindex, column=Var + 1).border = thin_border
                rowindex += 1
            Ws.sheet_view.selection[0].activeCell = 'A1'
        if sheet == "Sheet":
            Wb.remove(Wb[sheet])
        Wb.save(OutPath)
        Wb.close()

    # 生成中间件中的Messages sheet 并填充报文信息
    def GeneratorMessagesSheet(self, OutPath='', wsname=''):
        MessageNameListTxTemp = self.TxMessageNameList  # TxMessageNameList中已经赋值了Rx报文名
        MessageNameListRxTemp = self.RxMessageNameList #TxMessageNameList中已经赋值了Rx报文名
        MessageNameListTemp = MessageNameListRxTemp+MessageNameListTxTemp
        TxRxType='Rx'
        Msgnum=0
        for msg in MessageNameListTemp:
            if msg in MessageNameListRxTemp:
                TxRxType = 'Rx'
            else:
                TxRxType = 'Tx'
            txt = self.GetMessageInfo(msg, TxRxType, Msgnum)
            self.Messagesdf.loc[Msgnum] = ['ON']+['ON']+['']+txt[1:]
            Msgnum = Msgnum + 1
        MessagesdfTemp = self.Messagesdf
        MessagesdfTemp.loc[Msgnum] = ""  # add one space row between cyclic none 0 and cyclic 0
        self.Messagesdf = self.DataFrameCyclicSort(MessagesdfTemp, Msgnum)
        #insert ID columns to dataframe
        idx = 1
        for dfIdx in self.Messagesdf.index:
            if self.Messagesdf.loc[dfIdx, "Cyclic Time"] == '':
                continue
            self.Messagesdf.loc[dfIdx,'ID'] = idx
            idx += 1
        isWsExist = False
        isClear = True
        if os.path.exists(OutPath) == True:
            Wb = openpyxl.load_workbook(OutPath)
        else:
            Wb = Workbook()  # 创建新的excel

        for sheet in Wb.sheetnames:
            if sheet == wsname:
                Ws = Wb[wsname]
                rowindex = Ws.max_row + 1
                if isClear == True:
                    Wb.remove(Wb[wsname])
                    isWsExist = False
                else:
                    isWsExist = True
                break
            else:
                continue

        if not isWsExist:
            if wsname == 'Messages':
                Ws = Wb.create_sheet(wsname, len(Wb.worksheets) + 1)
                # Message的格式
                #Ws.row_dimensions[1].height = 11  # 行高设置，行1 40
                Ws.column_dimensions['A'].width = 12  # 列宽设置， 列A 10
                Ws.column_dimensions['B'].width = 12
                Ws.column_dimensions['C'].width = 12
                Ws.column_dimensions['D'].width = 12
                Ws.column_dimensions['E'].width = 24
                Ws.column_dimensions['F'].width = 12
                Ws.column_dimensions['G'].width = 12
                Ws.column_dimensions['H'].width = 12
            rowindex=1
            Var=1
            for Title in self.Messagesdf.columns:
                Ws.cell(row=rowindex, column=Var).value = Title
                Ws.cell(row=rowindex, column=Var).border = thin_border
                Ws.cell(row=rowindex, column=Var).font = Font(bold=True)
                Var += 1
            rowindex += 1
            for id in self.Messagesdf.index:
                for Var in range(0, len(self.Messagesdf.columns)):
                    Ws.cell(row=rowindex, column=Var + 1).value = self.Messagesdf.loc[id][Var]
                    Ws.cell(row=rowindex, column=Var + 1).border = thin_border
                rowindex += 1
            Ws.sheet_view.selection[0].activeCell = 'A1'
        if sheet == "Sheet":
            Wb.remove(Wb[sheet])
        Wb.save(OutPath)
        Wb.close()

    # 生成中间件中空的Rx_Signals sheet
    def GeneratorRxSignalsSheet(self, OutPath='', wsname=''):
        isWsExist = False
        isClear = True
        if os.path.exists(OutPath) == True:
            Wb = openpyxl.load_workbook(OutPath)
        else:
            Wb = Workbook()  # 创建新的excel

        for sheet in Wb.sheetnames:
            if sheet == wsname:
                Ws = Wb[wsname]
                rowindex = Ws.max_row + 1
                if isClear == True:
                    Wb.remove(Wb[wsname])
                    isWsExist = False
                else:
                    isWsExist = True
                break
            else:
                continue

        if not isWsExist:
            if wsname == 'RX_Signals':
                Ws = Wb.create_sheet(wsname, len(Wb.worksheets) + 1)
                # Message的格式
                Ws.row_dimensions[1].height = 40  # 行高设置，行1 40
                Ws.column_dimensions['A'].width = 12  # 列宽设置， 列A 10
                Ws.column_dimensions['B'].width = 12
                Ws.column_dimensions['C'].width = 12
                Ws.column_dimensions['D'].width = 12
                Ws.column_dimensions['E'].width = 12
                Ws.column_dimensions['F'].width = 12
                Ws.column_dimensions['G'].width = 12
                Ws.column_dimensions['H'].width = 12
        rowindex = 1
        Var = 1
        for Title in self.RxSignaldf.columns:
            Ws.cell(row=rowindex, column=Var).value = Title
            Ws.cell(row=rowindex, column=Var).border = thin_border
            Ws.cell(row=rowindex, column=Var).font = Font(bold=True)
            Ws.cell(row=rowindex, column=Var).alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            Var += 1
        tmp = len(self.RxSignaldf.columns)
        Ws.merge_cells(start_row=rowindex, end_row=rowindex, start_column=tmp, end_column=tmp+10)
        if sheet == "Sheet":
            Wb.remove(Wb[sheet])
        Wb.save(OutPath)
        Wb.close()

    # 生成中间件中空的Tx_Signals sheet
    def GeneratorTxSignalsSheet(self, OutPath='', wsname=''):
        isWsExist = False
        isClear = True
        if os.path.exists(OutPath) == True:
            Wb = openpyxl.load_workbook(OutPath)
        else:
            Wb = Workbook()  # 创建新的excel

        for sheet in Wb.sheetnames:
            if sheet == wsname:
                Ws = Wb[wsname]
                rowindex = Ws.max_row + 1
                if isClear == True:
                    Wb.remove(Wb[wsname])
                    isWsExist = False
                else:
                    isWsExist = True
                break
            else:
                continue

        if not isWsExist:
            if wsname == 'TX_Signals':
                Ws = Wb.create_sheet(wsname, len(Wb.worksheets) + 1)
                # Message的格式
                Ws.row_dimensions[1].height = 40  # 行高设置，行1 40
                Ws.column_dimensions['A'].width = 12  # 列宽设置， 列A 10
                Ws.column_dimensions['B'].width = 12
                Ws.column_dimensions['C'].width = 12
                Ws.column_dimensions['D'].width = 12
                Ws.column_dimensions['E'].width = 12
                Ws.column_dimensions['F'].width = 12
                Ws.column_dimensions['G'].width = 12
                Ws.column_dimensions['H'].width = 12
                Ws.column_dimensions['T'].width = 18
        rowindex = 1
        Var = 1
        for Title in self.TxSignaldf.columns:
            Ws.cell(row=rowindex, column=Var).value = Title
            Ws.cell(row=rowindex, column=Var).border = thin_border
            Ws.cell(row=rowindex, column=Var).font = Font(bold=True)
            Ws.cell(row=rowindex, column=Var).alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            Var += 1
        #tmp = len(self.TxSignaldf.columns)
        #Ws.merge_cells(start_row=rowindex, end_row=rowindex, start_column=tmp, end_column=tmp+10)
        if sheet == "Sheet":
            Wb.remove(Wb[sheet])
        Wb.save(OutPath)
        Wb.close()

    #生成中间件中空的Develop_Signals sheet
    def GeneratorDevelopSignalsSheet(self, OutPath='', wsname=''):
        isWsExist = False
        isClear = True
        if os.path.exists(OutPath) == True:
            Wb = openpyxl.load_workbook(OutPath)
        else:
            Wb = Workbook()  # 创建新的excel

        for sheet in Wb.sheetnames:
            if sheet == wsname:
                Ws = Wb[wsname]
                rowindex = Ws.max_row + 1
                if isClear == True:
                    Wb.remove(Wb[wsname])
                    isWsExist = False
                else:
                    isWsExist = True
                break
            else:
                continue

        if not isWsExist:
            if wsname == 'Develop_Signals':
                Ws = Wb.create_sheet(wsname, len(Wb.worksheets) + 1)
                # Message的格式
                Ws.row_dimensions[1].height = 40  # 行高设置，行1 40
                Ws.column_dimensions['A'].width = 12  # 列宽设置， 列A 10
                Ws.column_dimensions['B'].width = 12
                Ws.column_dimensions['C'].width = 12
                Ws.column_dimensions['D'].width = 12
                Ws.column_dimensions['E'].width = 12
                Ws.column_dimensions['F'].width = 12
                Ws.column_dimensions['G'].width = 12
                Ws.column_dimensions['H'].width = 12
                Ws.column_dimensions['V'].width = 16
        rowindex = 1
        Var = 1
        for Title in self.DevelopSignalsdf.columns:
            Ws.cell(row=rowindex, column=Var).value = Title
            Ws.cell(row=rowindex, column=Var).border = thin_border
            Ws.cell(row=rowindex, column=Var).font = Font(bold=True)
            Ws.cell(row=rowindex, column=Var).alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            Var += 1
        #tmp = len(self.TxSignaldf.columns)
        #Ws.merge_cells(start_row=rowindex, end_row=rowindex, start_column=tmp, end_column=tmp+10)
        if sheet == "Sheet":
            Wb.remove(Wb[sheet])
        Wb.save(OutPath)
        Wb.close()

    # 生成中间件中空的Check_Result sheet
    def GeneratorCheckResultSheet(self, OutPath='', wsname=''):
        isWsExist = False
        isClear = True
        if os.path.exists(OutPath) == True:
            Wb = openpyxl.load_workbook(OutPath)
        else:
            Wb = Workbook()  # 创建新的excel

        for sheet in Wb.sheetnames:
            if sheet == wsname:
                Ws = Wb[wsname]
                rowindex = Ws.max_row + 1
                if isClear == True:
                    Wb.remove(Wb[wsname])
                    isWsExist = False
                else:
                    isWsExist = True
                break
            else:
                continue

        if not isWsExist:
            if wsname == 'Check_Result':
                Ws = Wb.create_sheet(wsname, len(Wb.worksheets) + 1)
                # Message的格式
                #Ws.row_dimensions[1].height = 40  # 行高设置，行1 40
                Ws.column_dimensions['A'].width = 12  # 列宽设置， 列A 10
                Ws.column_dimensions['B'].width = 80
                Ws.column_dimensions['C'].width = 12
                Ws.column_dimensions['D'].width = 12
                Ws.column_dimensions['E'].width = 12
                Ws.column_dimensions['F'].width = 12
                Ws.column_dimensions['G'].width = 12
                Ws.column_dimensions['H'].width = 12
        rowindex = 1
        Var = 1
        for Title in self.CheckResultdf.columns:
            Ws.cell(row=rowindex, column=Var).value = Title
            Ws.cell(row=rowindex, column=Var).border = thin_border
            Ws.cell(row=rowindex, column=Var).font = Font(bold=True)
            Ws.cell(row=rowindex, column=Var).alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            Var += 1
        #tmp = len(self.TxSignaldf.columns)
        #Ws.merge_cells(start_row=rowindex, end_row=rowindex, start_column=tmp, end_column=tmp+10)
        if sheet == "Sheet":
            Wb.remove(Wb[sheet])
        Wb.save(OutPath)
        Wb.close()

    # making data frame from xlsx file to data frame
    # 生成AllDBC要从xlsx文件中读取信息到data frame中
    def ReadOutputSheetAsDataFrame(self, InputPath='', wsname=''):
        WsDf = pd.read_excel(io=InputPath, sheet_name=wsname)
        return WsDf

    #给excel写title 这个函数没有很清晰，不知道要干嘛
    def WriteTitleToExcel(self, OutPath, wsname, df, colPos, isClearSheet):
        isWsExist = False
        isClear = isClearSheet
        if os.path.exists(OutPath) == True:
            Wb = openpyxl.load_workbook(OutPath)
        else:
            Wb = Workbook()  # 创建新的excel
        for sheet in Wb.sheetnames:
            if sheet == wsname:
                Ws = Wb[wsname]
                rowindex = Ws.max_row + 1
                if isClear == True:
                    Wb.remove(Wb[wsname])
                    isWsExist = False
                else:
                    isWsExist = True
                break
            else:
                continue
        if isWsExist == False:
            Ws = Wb.create_sheet(wsname, len(Wb.worksheets) + 1)
            Ws.column_dimensions['C'].width = 16
            Ws.column_dimensions['B'].width = 28
            Ws.column_dimensions['M'].width = 28
            Ws.column_dimensions['O'].width = 39
            Ws.column_dimensions['N'].width = 14
            Ws.column_dimensions['AE'].width = 52
            Ws.column_dimensions['AX'].width = 52
            Ws.column_dimensions['AZ'].width = 10
        rowIdx=1
        colIdx=1
        for Title in df.columns:
            Ws.cell(row=rowIdx, column=colPos + colIdx).value = Title
            Ws.cell(row=rowIdx, column=colPos + colIdx).border = thin_border
            colIdx += 1
        if sheet == "Sheet":
            Wb.remove(Wb[sheet])
        Wb.save(OutPath)
        Wb.close()


    #修改信号的类型 比如Unsign 32 就是Uint32
    def SignalNameCombine(self, DBCValueType, DBCSignalStorageBits, SignalName, ID):
        sigName = 'g_'
        if DBCValueType == 'Unsigned':
            sigName += 'u'
        else:
            sigName += 's'
        sigName += str(DBCSignalStorageBits) + '_'
        sigName += SignalName + '_'
        sigName += ID + '_'
        sigName += '0'
        return sigName

    #获取DBC信号存储位 >32取32 ， 且要标红
    def GetDBCSignalStorageBits(self, df, idx):
        StorageBit = 0
        signalLength = df.loc[idx]['Length']
        if signalLength > 32:
            StorageBit = 32
            self.Exceed32bitStorage.append(self.GreenRowIdx)
        elif signalLength > 16:
            StorageBit = 32
        elif signalLength > 8:
            StorageBit = 16
        else:
            StorageBit = 8
        return StorageBit

    #通过中间件开发人员手填的报文类型来录入AllDBC
    def GetFrameType(self, Frame, MsgType):
        FrameType = ''
        if Frame == 'Application':
            FrameType = 'COM_NORMAL'
        elif Frame == 'NM':
            FrameType = 'COM_NET'
        elif Frame == 'DIAG_Fun':
            FrameType = 'DIAG_Fun'
        elif Frame == 'DIAG_Phy':
            FrameType = 'DIAG_Phy'
        elif Frame == 'DIAG_TX':
            FrameType = 'DIAG'
        elif Frame == 'DIAG':
            FrameType = 'None'
        elif Frame == 'XCP':
            if MsgType == 'Rx':
                FrameType = 'XCP_RX'
            else:
                FrameType = 'XCP_TX'
        elif Frame == 'Develop':
            FrameType = 'COM_Develop'
        elif Frame == 'EOL':
            if MsgType == 'Rx':
                FrameType = 'EOL_RX'
            else:
                FrameType = 'EOL_TX'
        else:
            FrameType = ''
        return FrameType

if __name__ == "__main__":
    r = PrintDbcToDoos()
    r.RunW()
